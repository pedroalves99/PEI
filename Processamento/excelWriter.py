import xlsxwriter
from openpyxl import load_workbook
import openpyxl
from copy import copy

def create_excel(filename, evalType):

    workbook = xlsxwriter.Workbook(filename)


    worksheet = workbook.add_worksheet(evalType)


    cell_format1 = workbook.add_format({ #blue
        'bold': True,
        'fg_color': '#20B2AA',
        'align': 'center',
        'border': 1,
        'valign': 'vcenter'})

    cell_format2 = workbook.add_format({ #green
        'bold': True,
        'fg_color': '#90EE90',
        'align': 'center',
        'border': 1,
        'valign': 'vcenter'})


    merge_format = workbook.add_format({ #laranja
        'bold': 1,
        'align': 'center',
        'border': 1,
        'valign': 'vcenter',
        'fg_color': '#FFA07A'})

    cell_format3 = workbook.add_format({  # rosa
        'bold': True,
        'fg_color': '#F0E68C',
        'align': 'center',
        'border': 1,
        'valign': 'vcenter'})

    cell_format4 = workbook.add_format({
        'bold': True,
        'fg_color':'#FFDAB9',
        'align': 'center',
        'border': 1,
        'valign': 'vcenter'})

    cell_format5 = workbook.add_format({
        'bold': True,
        'fg_color': '#808000',
        'align': 'center',
        'border': 1,
        'valign': 'vcenter'})

    cell_format6 = workbook.add_format({
        'bold': True,
        'fg_color': '#F5F5DC',
        'align': 'center',
        'border': 1,
        'valign': 'vcenter'})

    worksheet.set_row(0, 25)
    worksheet.set_column(0, 0, 60)
    worksheet.set_column(1, 1, 15)
    #cardinal columns
    for i in range(2,10):
        worksheet.set_column(i, i, 5)

    #cardinal columns
    for i in range(10,18):
        worksheet.set_column(i, i, 5)


    for i in range(18, 26):
        worksheet.set_column(i, i, 5)

    for i in range(26,34):
        worksheet.set_column(i, i, 5)

    for i in range(34, 38):
        worksheet.set_column(i, i, 6)

    for i in range(38, 41):
        worksheet.set_column(i, i, 14)

    worksheet.merge_range('C1:J1', 'Global Contour Displacement(mm)', merge_format) #global vector

    worksheet.merge_range('K1:R1', 'Reference Contour Displacement(mm)', cell_format3)  # vectors %

    worksheet.merge_range('S1:Z1', 'Center of Mass Displacement(mm)', cell_format4) #centro de massa

    worksheet.merge_range('AA1:AH1', 'Reference Center of Mass Displacement(mm)', cell_format5)  #reference contour

    worksheet.merge_range('AI1:AL1', 'Displacement in x,y (mm) ', merge_format)  # reference contour


    worksheet.write('AM1', 'Area(mm^2)', cell_format6)  # Area contour
    worksheet.write('AM2', '', cell_format6)


    worksheet.write('AN1', 'Distance1(mm)', cell_format1)
    worksheet.write('AN2', '', cell_format1)

    worksheet.write('AO1', 'Distance2(mm)', cell_format2)  # Perimeter contour
    worksheet.write('AO2', '', cell_format2)


    worksheet.write("A1", "Video Name", cell_format1 )
    worksheet.write("A2", "", cell_format1)

    worksheet.write("B1", "Frame Number", cell_format2)
    worksheet.write("B2", "", cell_format2)

    ## cardinal global vectors
    worksheet.write("C2", "N", merge_format)
    worksheet.write("D2", "NE", merge_format)
    worksheet.write("E2", "E", merge_format)
    worksheet.write("F2", "SE", merge_format)
    worksheet.write("G2", "S", merge_format)
    worksheet.write("H2", "SW", merge_format)
    worksheet.write("I2", "W", merge_format)
    worksheet.write("J2", "NW", merge_format)

    ## cardinal centro de massa
    worksheet.write("K2", "N", cell_format3)
    worksheet.write("L2", "NE", cell_format3)
    worksheet.write("M2", "E", cell_format3)
    worksheet.write("N2", "SE", cell_format3)
    worksheet.write("O2", "S", cell_format3)
    worksheet.write("P2", "SW", cell_format3)
    worksheet.write("Q2", "W", cell_format3)
    worksheet.write("R2", "NW", cell_format3)

    ## cardinal reference contour
    worksheet.write("S2", "N", cell_format4)
    worksheet.write("T2", "NE", cell_format4)
    worksheet.write("U2", "E", cell_format4)
    worksheet.write("V2", "SE", cell_format4)
    worksheet.write("W2", "S", cell_format4)
    worksheet.write("X2", "SW", cell_format4)
    worksheet.write("Y2", "W", cell_format4)
    worksheet.write("Z2", "NW", cell_format4)

    worksheet.write("AA2", "N", cell_format5)
    worksheet.write("AB2", "NE", cell_format5)
    worksheet.write("AC2", "E", cell_format5)
    worksheet.write("AD2", "SE", cell_format5)
    worksheet.write("AE2", "S", cell_format5)
    worksheet.write("AF2", "SW", cell_format5)
    worksheet.write("AG2", "W", cell_format5)
    worksheet.write("AH2", "NW", cell_format5)

    worksheet.write("AI2", "x", merge_format)
    worksheet.write("AJ2", "xRef", merge_format)
    worksheet.write("AK2", "y", merge_format)
    worksheet.write("AL2", "yRef", merge_format)

    xlsxwriter.Workbook(filename)
    workbook.close()

def _copy_dimensions():
    global sheet1, sheet
    i = 0

    src = getattr(sheet, 'column_dimensions')
    target = getattr(sheet1, 'column_dimensions')
    for key, dim in src.items():
        i+=1
        target[key] = copy(dim)
        target[key].worksheet = sheet1
        if i == 41:
            break;

def copy_header(filename, evalType):
    global sheet1, sheet
    workbook = load_workbook(filename);

    index = len(workbook.get_sheet_names())
    workbook.create_sheet(evalType, index)

    sheet1  = workbook.worksheets[index]
    sheet = workbook.worksheets[0]

    _copy_dimensions()

    for i in range(1,3):
        for j in range(1,42):
            sheet1.cell(i,j).value = workbook.worksheets[0].cell(i,j).value
            sheet1.cell(i, j)._style = copy(workbook.worksheets[0].cell(i, j)._style)

    sheet1.merge_cells('C1:J1')  # global vector

    sheet1.merge_cells('K1:R1')  # centro de massa

    sheet1.merge_cells('S1:Z1')  # reference contour

    sheet1.merge_cells('AA1:AH1')

    sheet1.merge_cells('AI1:AL1')

    sheet1.row_dimensions[1].height = 25

    workbook.save(filename)

def get_style(ws, index, y):
    if ws.cell(index - 1, y).has_style:
        print("ll")
        ws.cell(index, y)._style = copy(ws.cell(index - 1, y)._style)
    else: print("não encontrou o estilo")

def set_value(ws, index, p, value):
    get_style(ws, index, p)
    ws.cell(index, p).value = value


def add_data(filename, video_name, frame_num, global_dislocation, reference_dislocation, cm_dislocation, cm_ref_dis, x, xref, y, yref, area, distance1, distance2, index):


    workbook = load_workbook(filename)

    ws = workbook.worksheets[len(workbook.get_sheet_names()) - 1]


    print(index)
    # passa o style para a célula de baixo

    set_value(ws, index, 1, video_name)

    set_value(ws, index, 2, frame_num)

    p = 3

    if len(global_dislocation) == 8:
        for elem in global_dislocation:
            set_value(ws, index, p, round(elem, 3))
            p+=1

    else:
        max8 = p + 8
        for v in range (p, max8):
            set_value(ws, index, p, "")
            p+=1

    if len(reference_dislocation) == 8:
        for elem in reference_dislocation:
            set_value(ws, index, p, round(elem, 3))
            p += 1
    else:
        max8 = p + 8
        for v in range(p, max8):
            set_value(ws, index, p, "")
            p += 1

    if len(cm_dislocation) == 8:
        for elem in cm_dislocation:
            set_value(ws, index, p, round(elem, 3))
            p += 1
    else:
        max8 = p + 8
        for v in range(p, max8):
            set_value(ws, index, p, "")
            p += 1

    if len(cm_ref_dis) == 8:
        for elem in reference_dislocation:
            set_value(ws, index, p, round(elem, 3))
            p += 1
    else:
        max8 = p + 8
        for v in range(p, max8):
            set_value(ws, index, p, "")
            p += 1


    set_value(ws, index, p, x)
    p += 1


    set_value(ws, index, p, xref)
    p += 1

    set_value(ws, index, p, y)
    p += 1

    set_value(ws, index, p, yref)
    p += 1

    set_value(ws, index, p, area)
    p += 1

    set_value(ws, index, p, distance1)
    p += 1

    set_value(ws, index, p, distance2)
    p += 1



    workbook.save(filename)




